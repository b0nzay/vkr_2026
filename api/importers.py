from __future__ import annotations

import io
import json
import posixpath
import zipfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import PurePosixPath

from django.core.files.base import ContentFile
from django.db import transaction

from catalog.models import (
    BodyType,
    Brand,
    CarModel,
    Category,
    Generation,
    Product,
    ProductBodyTypeCompatibility,
    ProductTechVariantCompatibility,
    TechVariant,
)

from .import_schemas import (
    CONFLICT_MODE_SKIP,
    CONFLICT_MODE_STOP,
    CONFLICT_MODE_UPDATE,
    FIELD_ALIASES,
    REQUIRED_FIELDS,
    SUPPORTED_CONFLICT_MODES,
    SUPPORTED_ENTITY_TYPES,
    SUPPORTED_IMAGE_EXTENSIONS,
)

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


class ImportValidationError(Exception):
    pass


@dataclass
class RowError:
    row: int
    field: str
    message: str


class ImagesArchive:
    def __init__(self, upload_file):
        self._files: dict[str, bytes] = {}
        self._invalid_paths: list[str] = []
        if not upload_file:
            return
        self._load(upload_file)

    def _load(self, upload_file):
        try:
            upload_file.seek(0)
        except Exception:
            pass
        data = upload_file.read()
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                normalized = self.normalize_path(info.filename)
                if not normalized:
                    self._invalid_paths.append(info.filename)
                    continue
                self._files[normalized] = zf.read(info.filename)

    @staticmethod
    def normalize_path(raw_path: str | None) -> str | None:
        if raw_path is None:
            return None
        p = str(raw_path).strip().replace('\\', '/')
        if not p:
            return None
        if p.startswith('/'):
            return None
        normalized = posixpath.normpath(p)
        if normalized.startswith('../') or normalized == '..':
            return None
        if ':' in normalized:
            return None
        return normalized

    def get_content_file(self, image_path: str | None) -> ContentFile | None:
        normalized = self.normalize_path(image_path)
        if not normalized:
            return None
        content = self._files.get(normalized)
        if content is None:
            return None
        suffix = PurePosixPath(normalized).suffix.lower()
        if suffix and suffix not in SUPPORTED_IMAGE_EXTENSIONS:
            raise ImportValidationError(
                f'Недопустимое расширение файла "{suffix}" для image_path="{normalized}".'
            )
        return ContentFile(content, name=PurePosixPath(normalized).name)

    @property
    def invalid_paths(self) -> list[str]:
        return list(self._invalid_paths)


def _normalize_key(value: str) -> str:
    return str(value).strip().lower().replace(' ', '_')


def _clean_str(v) -> str:
    if v is None:
        return ''
    return str(v).strip()


def _normalize_body_type(raw: str) -> str:
    value = _clean_str(raw).upper().replace(' ', '_').replace('-', '_')
    if not value:
        return value
    if value in BodyType.Types.values:
        return value
    display_map = {label.upper().replace(' ', '_'): code for code, label in BodyType.Types.choices}
    return display_map.get(value, value)


def _normalize_transmission(raw: str) -> str:
    value = _clean_str(raw).upper()
    if value in TechVariant.TransmissionType.values:
        return value
    display_map = {label.upper(): code for code, label in TechVariant.TransmissionType.choices}
    return display_map.get(value, value)


def _normalize_engine_type(raw: str) -> str:
    value = _clean_str(raw).upper()
    if not value:
        return TechVariant.EngineType.NATURALLY_ASPIRATED
    if value in TechVariant.EngineType.values:
        return value
    display_map = {label.upper(): code for code, label in TechVariant.EngineType.choices}
    return display_map.get(value, value)


def _normalize_fuel_type(raw: str) -> str:
    value = _clean_str(raw).upper()
    if not value:
        return TechVariant.FuelType.PETROL
    if value in TechVariant.FuelType.values:
        return value
    display_map = {label.upper(): code for code, label in TechVariant.FuelType.choices}
    return display_map.get(value, value)


def _as_int(value, *, field: str):
    if value in (None, ''):
        return None
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        raise ImportValidationError(f'Поле "{field}" должно быть целым числом.')


def _as_decimal(value, *, field: str):
    if value in (None, ''):
        return None
    try:
        return Decimal(str(value).strip().replace(',', '.'))
    except (InvalidOperation, ValueError):
        raise ImportValidationError(f'Поле "{field}" должно быть числом.')


def _parse_refs(raw_value: str) -> list[str]:
    if raw_value in (None, ''):
        return []
    return [chunk.strip() for chunk in str(raw_value).split(';') if chunk.strip()]


def _extract_rows(entity_type: str, data_file):
    if entity_type not in SUPPORTED_ENTITY_TYPES:
        raise ImportValidationError('Некорректный entity_type.')
    if not data_file:
        raise ImportValidationError('Файл данных обязателен.')

    filename = str(getattr(data_file, 'name', '')).lower()
    if filename.endswith('.json'):
        payload = json.load(data_file)
        if not isinstance(payload, list):
            raise ImportValidationError('JSON должен содержать массив объектов.')
        return payload

    if filename.endswith('.xlsx'):
        if load_workbook is None:
            raise ImportValidationError('Для импорта XLSX требуется библиотека openpyxl.')
        wb = load_workbook(data_file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        result = []
        for row in rows[1:]:
            entry = {}
            for idx, cell in enumerate(row):
                if idx >= len(headers):
                    continue
                key = headers[idx]
                if not key:
                    continue
                entry[key] = cell
            if any(v not in (None, '') for v in entry.values()):
                result.append(entry)
        return result

    raise ImportValidationError('Поддерживаются только файлы .json и .xlsx')


def _normalize_row(entity_type: str, raw: dict):
    aliases = FIELD_ALIASES[entity_type]
    keymap = {_normalize_key(k): v for k, v in (raw or {}).items()}
    normalized: dict[str, object] = {}
    for field, field_aliases in aliases.items():
        for alias in field_aliases:
            k = _normalize_key(alias)
            if k in keymap:
                normalized[field] = keymap[k]
                break
    return normalized


def _validate_required(entity_type: str, row: dict):
    for field in REQUIRED_FIELDS[entity_type]:
        if _clean_str(row.get(field)) == '':
            raise ImportValidationError(f'Обязательное поле "{field}" не заполнено.')


class ImportEngine:
    def __init__(self, *, entity_type: str, conflict_mode: str, dry_run: bool, images_archive: ImagesArchive):
        self.entity_type = entity_type
        self.conflict_mode = conflict_mode
        self.dry_run = dry_run
        self.images_archive = images_archive

        self.created = 0
        self.updated = 0
        self.skipped = 0
        self.failed = 0
        self.conflicts = 0
        self.images_attached = 0
        self.images_missing = 0
        self.row_errors: list[RowError] = []

    def run(self, raw_rows: list[dict]):
        if self.conflict_mode not in SUPPORTED_CONFLICT_MODES:
            raise ImportValidationError('Некорректный conflict_mode.')

        with transaction.atomic():
            for idx, raw in enumerate(raw_rows, start=2):
                try:
                    row = _normalize_row(self.entity_type, raw)
                    _validate_required(self.entity_type, row)
                    if self.entity_type == 'cars':
                        self._import_car_row(idx, row)
                    elif self.entity_type == 'tech_variants':
                        self._import_tech_row(idx, row)
                    elif self.entity_type == 'products':
                        self._import_product_row(idx, row)
                    else:
                        raise ImportValidationError('Неизвестный тип сущности.')
                except Exception as exc:
                    self.failed += 1
                    self.row_errors.append(RowError(row=idx, field='row', message=str(exc)))
                    if self.conflict_mode == CONFLICT_MODE_STOP:
                        raise
            if self.dry_run:
                transaction.set_rollback(True)

        return {
            'total': len(raw_rows),
            'created': self.created,
            'updated': self.updated,
            'skipped': self.skipped,
            'failed': self.failed,
            'conflicts': self.conflicts,
            'images_attached': self.images_attached,
            'images_missing': self.images_missing,
            'errors': [{'row': e.row, 'field': e.field, 'message': e.message} for e in self.row_errors],
            'invalid_image_paths': self.images_archive.invalid_paths,
        }

    def _handle_conflict(self):
        self.conflicts += 1
        if self.conflict_mode == CONFLICT_MODE_SKIP:
            self.skipped += 1
            return 'skip'
        if self.conflict_mode == CONFLICT_MODE_STOP:
            raise ImportValidationError('Обнаружен конфликт ключа.')
        return 'update'

    def _attach_image(self, instance, field_name: str, image_path: str | None):
        if not image_path:
            return
        content = self.images_archive.get_content_file(image_path)
        if content is None:
            self.images_missing += 1
            if self.conflict_mode == CONFLICT_MODE_STOP:
                raise ImportValidationError(f'Файл изображения не найден: {image_path}')
            return
        getattr(instance, field_name).save(content.name, content, save=False)
        self.images_attached += 1

    def _import_car_row(self, row_num: int, row: dict):
        brand_name = _clean_str(row.get('brand'))
        model_name = _clean_str(row.get('model'))
        generation_name = _clean_str(row.get('generation'))
        body_type = _normalize_body_type(row.get('body_type'))
        if body_type not in BodyType.Types.values:
            raise ImportValidationError(f'Некорректный body_type: "{body_type}" (row {row_num}).')

        brand, _ = Brand.objects.get_or_create(name=brand_name)
        if _:
            self.created += 1
        model, _ = CarModel.objects.get_or_create(brand=brand, name=model_name)
        if _:
            self.created += 1
        generation, gen_created = Generation.objects.get_or_create(car_model=model, name=generation_name)
        if gen_created:
            self.created += 1

        body, body_created = BodyType.objects.get_or_create(generation=generation, name=body_type)
        if body_created:
            self.created += 1
        else:
            decision = self._handle_conflict()
            if decision == 'skip':
                return

        self._attach_image(brand, 'logo', _clean_str(row.get('brand_logo_path')))
        self._attach_image(generation, 'image', _clean_str(row.get('generation_image_path')))
        self._attach_image(body, 'image', _clean_str(row.get('body_type_image_path')))
        if not self.dry_run:
            brand.save()
            generation.save()
            body.save()
        if not body_created:
            self.updated += 1

    def _resolve_generation(self, row: dict):
        brand_name = _clean_str(row.get('brand'))
        model_name = _clean_str(row.get('model'))
        generation_name = _clean_str(row.get('generation'))
        generation = Generation.objects.filter(
            car_model__brand__name__iexact=brand_name,
            car_model__name__iexact=model_name,
            name__iexact=generation_name,
        ).first()
        if not generation:
            raise ImportValidationError(
                f'Поколение не найдено: brand={brand_name}, model={model_name}, generation={generation_name}.'
            )
        return generation

    def _import_tech_row(self, row_num: int, row: dict):
        generation = self._resolve_generation(row)
        engine_code = _clean_str(row.get('engine_code'))
        transmission_code = _clean_str(row.get('transmission_code'))
        transmission_type = _normalize_transmission(row.get('transmission_type'))
        if transmission_type not in TechVariant.TransmissionType.values:
            raise ImportValidationError(
                f'Некорректный transmission_type: "{transmission_type}" (row {row_num}).'
            )
        defaults = {
            'engine_type': _normalize_engine_type(row.get('engine_type')),
            'fuel_type': _normalize_fuel_type(row.get('fuel_type')),
            'gears': _as_int(row.get('gears'), field='gears'),
            'power_hp': _as_int(row.get('power_hp'), field='power_hp'),
            'torque_nm': _as_int(row.get('torque_nm'), field='torque_nm'),
            'notes': _clean_str(row.get('notes')),
            'transmission_type': transmission_type,
        }
        tech = TechVariant.objects.filter(
            generation=generation,
            engine_code=engine_code,
            transmission_code=transmission_code,
        ).first()
        if tech:
            decision = self._handle_conflict()
            if decision == 'skip':
                return
            for k, v in defaults.items():
                setattr(tech, k, v)
            if not self.dry_run:
                tech.save()
            self.updated += 1
            return

        tech = TechVariant(
            generation=generation,
            engine_code=engine_code,
            transmission_code=transmission_code,
            **defaults,
        )
        if not self.dry_run:
            tech.save()
        self.created += 1

    def _resolve_category(self, raw_category: str) -> Category:
        chunks = [c.strip() for c in str(raw_category).split('>') if c.strip()]
        if not chunks:
            chunks = [str(raw_category).strip()]
        parent = None
        current = None
        for chunk in chunks:
            if parent is not None and not getattr(parent, 'pk', None):
                current = None
            else:
                current = Category.objects.filter(name__iexact=chunk, parent=parent).first()
            if current is None:
                current = Category(name=chunk, parent=parent)
                if not self.dry_run:
                    current.save()
                self.created += 1
            parent = current
        return current

    def _resolve_body_type(self, ref: str):
        parts = [p.strip() for p in ref.split('|')]
        if len(parts) != 4:
            raise ImportValidationError('body_type_refs должен иметь формат "Brand|Model|Generation|BodyType".')
        brand_name, model_name, generation_name, body_type_name = parts
        body_type = BodyType.objects.filter(
            generation__car_model__brand__name__iexact=brand_name,
            generation__car_model__name__iexact=model_name,
            generation__name__iexact=generation_name,
            name=_normalize_body_type(body_type_name),
        ).first()
        if not body_type:
            raise ImportValidationError(f'Не найден body type ref: {ref}')
        return body_type

    def _resolve_tech_variant(self, ref: str):
        parts = [p.strip() for p in ref.split('|')]
        if len(parts) != 5:
            raise ImportValidationError(
                'tech_variant_refs должен иметь формат "Brand|Model|Generation|EngineCode|TransmissionCode".'
            )
        brand_name, model_name, generation_name, engine_code, transmission_code = parts
        tv = TechVariant.objects.filter(
            generation__car_model__brand__name__iexact=brand_name,
            generation__car_model__name__iexact=model_name,
            generation__name__iexact=generation_name,
            engine_code__iexact=engine_code,
            transmission_code__iexact=transmission_code,
        ).first()
        if not tv:
            raise ImportValidationError(f'Не найден tech variant ref: {ref}')
        return tv

    def _import_product_row(self, row_num: int, row: dict):
        sku = _clean_str(row.get('sku'))
        category = self._resolve_category(_clean_str(row.get('category')))
        price = _as_decimal(row.get('price'), field='price')
        if price is None or price <= 0:
            raise ImportValidationError(f'Некорректная цена в row {row_num}.')
        stock = _as_int(row.get('stock'), field='stock')
        if stock is None or stock < 0:
            raise ImportValidationError(f'Некорректный остаток в row {row_num}.')
        compatibility_mode = _clean_str(row.get('compatibility_mode')) or Product.CompatibilityMode.BODY_TYPE
        if compatibility_mode not in Product.CompatibilityMode.values:
            raise ImportValidationError(f'Некорректный compatibility_mode: "{compatibility_mode}".')

        defaults = {
            'name': _clean_str(row.get('name')),
            'price': price,
            'stock': stock,
            'description': _clean_str(row.get('description')),
            'category': category,
            'brand_name': _clean_str(row.get('brand_name')),
            'compatibility_mode': compatibility_mode,
        }
        product = Product.objects.filter(sku=sku).first()
        if product:
            decision = self._handle_conflict()
            if decision == 'skip':
                return
            for k, v in defaults.items():
                setattr(product, k, v)
            self._attach_image(product, 'image', _clean_str(row.get('image_path')))
            if not self.dry_run:
                product.save()
            self.updated += 1
        else:
            product = Product(sku=sku, **defaults)
            self._attach_image(product, 'image', _clean_str(row.get('image_path')))
            if not self.dry_run:
                product.save()
            self.created += 1

        if compatibility_mode == Product.CompatibilityMode.BODY_TYPE:
            refs = _parse_refs(row.get('body_type_refs'))
            if refs:
                if not self.dry_run:
                    ProductBodyTypeCompatibility.objects.filter(product=product).delete()
                for ref in refs:
                    body_type = self._resolve_body_type(ref)
                    if not self.dry_run:
                        ProductBodyTypeCompatibility.objects.get_or_create(product=product, body_type=body_type)
        if compatibility_mode == Product.CompatibilityMode.TECH_VARIANT:
            refs = _parse_refs(row.get('tech_variant_refs'))
            if refs:
                if not self.dry_run:
                    ProductTechVariantCompatibility.objects.filter(product=product).delete()
                for ref in refs:
                    tv = self._resolve_tech_variant(ref)
                    if not self.dry_run:
                        ProductTechVariantCompatibility.objects.get_or_create(product=product, tech_variant=tv)


def run_import(
    *,
    entity_type: str,
    data_file,
    conflict_mode: str = CONFLICT_MODE_STOP,
    dry_run: bool = True,
    images_zip=None,
):
    rows = _extract_rows(entity_type, data_file)
    images = ImagesArchive(images_zip)
    engine = ImportEngine(
        entity_type=entity_type,
        conflict_mode=conflict_mode,
        dry_run=dry_run,
        images_archive=images,
    )
    return engine.run(rows)

